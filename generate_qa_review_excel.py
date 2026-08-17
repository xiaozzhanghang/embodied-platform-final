import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()

# ==================== SHEET 1: 需求任务清单 ====================
ws1 = wb.active
ws1.title = "需求与任务清单"
ws1.views.sheetView[0].showGridLines = True

headers1 = ["模块", "功能菜单", "层级/页面", "需求类型", "功能名称", "任务描述与业务规则", "前端路由/位置", "验收标准"]
ws1.append(headers1)

data_items = [
    # ---------------- 数据质检 ----------------
    [
        "数据质检",
        "数据质检",
        "一级列表",
        "功能需求",
        "分包质检多维检索与页签分类",
        "支持按分包实例ID、任务名称、来源计划名称/ID、采集人员、指派质检员、分包送检时间等多维度组合检索，支持按【全部、待质检、质检中、已通过、未通过】状态分类页签进行实时数据筛选，各筛选项支持一键重置与清空。",
        "/collection/qa",
        "1. 筛选条件组合生效且重置正常；\n2. 页签数字与分包状态完全一致；\n3. 列表支持按送检时间倒序排列。"
    ],
    [
        "数据质检",
        "数据质检",
        "一级列表",
        "业务规则",
        "质检员指派与重新分配",
        "质检主管可在分包主表中选中单条分包点击【重新分配】，或勾选多个分包点击【批量分配】，弹出分配质检员弹窗，指定目标质检员后提交更新分包归属，流转状态自动保持并同步通知质检员。",
        "/collection/qa (列表操作列及顶部工具栏)",
        "1. 单条及批量分配弹窗均能正确更新质检员；\n2. 操作列按钮命名统一为【重新分配】。"
    ],
    [
        "数据质检",
        "数据质检",
        "一级列表",
        "交互规则",
        "主表操作列与工具栏标准规范",
        "质检主表操作列统一提供【进入】（下钻至分包质检详情）和【重新分配】按钮，去除非法物理删除按钮及去标注跨环节按钮；列表工具栏标配【刷新、表格密度（默认/中等/紧凑）、列设置（显示/隐藏列勾选+重置）】组件。",
        "/collection/qa (操作列及工具栏)",
        "1. 操作列仅保留【进入】与【重新分配】；\n2. 密度切换与列设置勾选即时生效。"
    ],
    [
        "数据质检",
        "数据质检详情",
        "二级详情页",
        "功能需求",
        "分包质检概览指标与状态过滤",
        "顶部展示分包基本信息及指标卡片：【总数据量、待质检、质检中、已通过、未通过】，下方支持按 Episode ID 模糊搜索及状态页签（待质检/质检中/已通过/未通过/全部）切换过滤。",
        "/collection/qa/[instanceId]",
        "1. 各状态 Episode 数量与卡片统计严格对应；\n2. 点击页签即时过滤下方数据表格。"
    ],
    [
        "数据质检",
        "数据质检详情",
        "二级详情页",
        "功能需求",
        "一键全部质检通过",
        "在数据列表右上角工具栏设置【一键全部通过】高亮绿色按钮，点击后弹出确认框提示待质检总数；确认后将本分包内全部未通过/待质检 Episode 数据一键变更为【已通过】状态，并给出全局成功反馈，数据自动流转至待标注合格池。",
        "/collection/qa/[instanceId] (表格工具栏)",
        "1. 按钮位于列表右上角工具栏，点击出二次确认；\n2. 确认后批量变更为【已通过】并弹出成功 Toast。"
    ],
    [
        "数据质检",
        "数据质检详情",
        "二级详情页",
        "业务规则",
        "质检不通过直接标记机制（无需重采）",
        "对于采集存在瑕疵或不合格的 Episode 数据，质检员可单条或批量标记为【未通过】标签（流程终止/缺陷数据归档），业务上不再走打回重采回路，确保已采集流水独立封存。",
        "/collection/qa/[instanceId]",
        "1. 勾选数据点击【批量标记未通过】后状态更新为未通过；\n2. 页面与文案中不再出现“打回重采”回路。"
    ],
    [
        "数据质检",
        "质检工作台",
        "三级作业台",
        "功能需求",
        "多视角视频与点云质检回放",
        "支持多路相机视频（头部、主臂、腕部等）多视角同步回放，支持 3D 点云渲染、时间轴拖拽与播放倍速切换；质检员快速核验动作是否完整并执行单条【通过】或【不通过】判定，判定后自动切换下一条。",
        "/annotation/audit/[instanceId]/[episodeId]?mode=audit",
        "1. 视频与点云时间轴同步播放无卡顿；\n2. 点击判定后自动跳转下一条 Episode。"
    ],

    # ---------------- 数据审核 ----------------
    [
        "数据审核",
        "数据审核",
        "一级列表",
        "功能需求",
        "独立数据审核任务看板与检索",
        "从标注工作台完全解耦的独立审核管理菜单，支持按任务ID、任务名称、关联资产源、标注类型（点/框/范围/混合）、审核状态、指派标注员、指派审核员联合检索与页签分类（全部/待分配/进行中/已完成/暂停）。",
        "/annotation/review",
        "1. 独立二级菜单【数据审核】入口正常；\n2. 移除新建标注任务等非审核功能入口；\n3. 多条件查询与页签切换正常。"
    ],
    [
        "数据审核",
        "数据审核",
        "一级列表",
        "业务规则",
        "审核员重新分配与操作规范",
        "支持审核主管对审核任务进行【重新分配】（单条或批量变更指定审核员）；操作列规范统一为【进入】（下钻至 Episode 审核明细）与【重新分配】，不提供物理删除按钮，删除由上游数据标注任务管理。",
        "/annotation/review (操作列)",
        "1. 操作列仅保留【重新分配】与【进入】；\n2. 重新分配提交后任务归属实时更新。"
    ],
    [
        "数据审核",
        "审核明细列表",
        "二级列表页",
        "功能需求",
        "Episode 审核明细与合格率统计",
        "展示分包内所有 Episode 明细数据，顶部实时计算【审核通过率（%）】、【已通过数】、【未通过(驳回重标)数】指标；支持按 Episode ID 检索与状态页签过滤，操作列统一提供【审核】（进入审核台）与【查看】（只读回放）。",
        "/annotation/review-list",
        "1. Episode ID 单行展示（已移除冗余副标题）；\n2. 合格率按公式（通过/(通过+驳回)）动态计算；\n3. 列表工具栏标配刷新、密度与列设置。"
    ],
    [
        "数据审核",
        "审核工作台",
        "三级作业台",
        "功能需求",
        "3D包围框与动作轨迹质量审核",
        "进入审核工作台，审核员可核验关键帧区间、3D 包围框贴合度、动作技能步骤标签是否正确，支持单条判定【通过】或【不通过】；审核通过的数据自动标记为已验收，流转至高质量具身数据集归档池。",
        "/annotation/audit/[instanceId]/[episodeId]?mode=audit",
        "1. 审核模式下显示通过/不通过操作条；\n2. 审核通过后状态变更为【已通过】并自动跳下一条。"
    ],
    [
        "数据审核",
        "审核工作台",
        "三级作业台",
        "业务规则",
        "审核不通过弹窗填写驳回理由与重新标注流转",
        "审核员点击【不通过】时，系统阻断并弹出【审核不通过 — 请输入驳回理由】弹窗；支持快捷选择不合格原因标签（关键帧时机不准、3D包围框不紧密、动作步骤标签选错、轨迹拟合偏移、文字描述不规范）并录入详细理由；确认后状态变更为【未通过/审核驳回】，携带驳回原因回退至标注工作台供标注员重新修改标注。",
        "/annotation/audit/[instanceId]/[episodeId] (不通过弹窗)",
        "1. 点击【不通过】必须弹出理由录入弹窗；\n2. 提交后数据流转回标注工作台，标注员可见驳回原因并重新标注。"
    ],
    [
        "数据审核",
        "标注工作台解耦",
        "系统架构",
        "架构规则",
        "标注工作台与数据审核彻底解耦",
        "标注工作台（/annotation/audit）定位为纯粹的标注员作业看板，已彻底移除【审核全部数据】、操作列【审核】按钮以及批量审核通过/驳回逻辑；所有审核验收行为全量收敛至【数据审核】菜单中。",
        "/annotation/audit 及 /annotation/audit/[id]",
        "1. 标注工作台操作列仅保留【标注】、【查看】、【重置】；\n2. 标注工作台内无任何审核权限与入口。"
    ]
]

for row in data_items:
    ws1.append(row)

# ==================== SHEET 2: 状态机与业务流转规范 ====================
ws2 = wb.create_sheet(title="状态流转与权限矩阵")
ws2.views.sheetView[0].showGridLines = True

headers2 = ["业务环节", "前置状态", "触发动作", "后置状态", "涉及角色", "界面位置", "流转说明与数据变更"]
ws2.append(headers2)

status_flow_data = [
    ["数据质检", "已送检 (待质检)", "质检员点击【通过】或【一键全部通过】", "质检已通过", "质检员 / 质检主管", "数据质检详情列表 / 质检工作台", "单条或全包数据合格，进入【待标注合格数据池】，供数据标注任务勾选创建。"],
    ["数据质检", "已送检 (待质检)", "质检员点击【不通过】或【批量标记未通过】", "质检未通过", "质检员", "数据质检详情列表 / 质检工作台", "直接打上【未通过】标签，缺陷归档，流程终止，不再打回重采。"],
    ["数据标注", "质检已通过", "PM 新建标注任务并指派人员", "未标注", "标注 PM", "数据标注 / 新建任务", "从合格数据池选取数据，指定标注员与审核员，生成标注任务实例并推送至标注工作台。"],
    ["标注工作台", "未标注", "标注员完成逐帧/3D标注并点击【提交】", "已标注 (待审核)", "标注员", "标注工作台", "标注作业完成，数据自动流转至【数据审核】菜单待审核池。"],
    ["数据审核", "已标注 (待审核)", "审核员核验合格并点击【通过】", "审核通过 (已归档)", "审核员 / 验收专家", "审核工作台", "标注质量验收合格，数据正式归档入库成为高质量具身算法训练集。"],
    ["数据审核", "已标注 (待审核)", "审核员点击【不通过】并填写驳回理由", "审核驳回 (待重新标注)", "审核员", "审核工作台 (驳回弹窗)", "附带驳回具体原因回退至【标注工作台】，标注员接收后重新修改标注并再次提交。"]
]

for row in status_flow_data:
    ws2.append(row)

# ==================== STYLING & FORMATTING ====================
header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
header_fill_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_fill_teal = PatternFill(start_color="006D75", end_color="006D75", fill_type="solid")

data_font = Font(name="微软雅黑", size=9)
bold_font = Font(name="微软雅黑", size=9, bold=True)

align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Style Sheet 1
for col_num in range(1, len(headers1) + 1):
    cell = ws1.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill_blue
    cell.alignment = align_center
    cell.border = thin_border
ws1.row_dimensions[1].height = 28

for row_num in range(2, len(data_items) + 2):
    ws1.row_dimensions[row_num].height = 56
    module_val = ws1.cell(row=row_num, column=1).value
    for col_num in range(1, len(headers1) + 1):
        cell = ws1.cell(row=row_num, column=col_num)
        cell.font = data_font
        cell.border = thin_border
        if col_num in [1, 2, 3, 4, 5]:
            cell.alignment = align_center
            if col_num == 1:
                cell.font = bold_font
                if module_val == "数据质检":
                    cell.fill = PatternFill(start_color="E6FFFB", end_color="E6FFFB", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F6FFED", end_color="F6FFED", fill_type="solid")
        elif col_num == 7:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

col_widths1 = {
    "A": 12,  # 模块
    "B": 14,  # 功能菜单
    "C": 14,  # 层级/页面
    "D": 12,  # 需求类型
    "E": 24,  # 功能名称
    "F": 46,  # 任务描述与业务规则
    "G": 24,  # 前端路由/位置
    "H": 36   # 验收标准
}
for col_letter, width in col_widths1.items():
    ws1.column_dimensions[col_letter].width = width

# Style Sheet 2
for col_num in range(1, len(headers2) + 1):
    cell = ws2.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill_teal
    cell.alignment = align_center
    cell.border = thin_border
ws2.row_dimensions[1].height = 28

for row_num in range(2, len(status_flow_data) + 2):
    ws2.row_dimensions[row_num].height = 42
    for col_num in range(1, len(headers2) + 1):
        cell = ws2.cell(row=row_num, column=col_num)
        cell.font = data_font
        cell.border = thin_border
        if col_num in [1, 2, 4, 5]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

col_widths2 = {
    "A": 14,  # 业务环节
    "B": 18,  # 前置状态
    "C": 26,  # 触发动作
    "D": 22,  # 后置状态
    "E": 18,  # 涉及角色
    "F": 28,  # 界面位置
    "G": 46   # 流转说明与数据变更
}
for col_letter, width in col_widths2.items():
    ws2.column_dimensions[col_letter].width = width

output_file = "/Users/zhangxiaozhang/Desktop/具身智能平台原型/0714/prototype/具身智能平台_数据质检与数据审核_需求任务表.xlsx"
wb.save(output_file)
print(f"Excel successfully created at: {output_file}")
