import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "需求与任务清单"

# Ensure gridlines are visible
ws.views.sheetView[0].showGridLines = True

# Headers
headers = ["版本", "类型", "菜单", "任务描述"]
ws.append(headers)

# Data rows based on project features
data = [
    # 任务派发
    ["V1.0", "功能需求", "任务派发", "任务中心多维度检索与过滤：支持按一级项目、二级项目、任务书、任务名称、任务ID、创建人、场景分类、任务用途、采集模式、遥操类型及设备类型联合筛选，各筛选项支持一键清空（allowClear）及重置还原初始态。"],
    ["V1.0", "功能需求", "任务派发", "任务状态分类页签：支持按“全部”、“进行中”、“排队中”、“已完成”页签快速切换并过滤任务列表。"],
    ["V1.0", "功能需求", "任务派发", "批量添加标注与标注流绑定：支持勾选任务后点击“批量添加标注”，选择点标注、范围标注、框标注及混合标注模式，并动态渲染对应质检员、标注员、审核员及“自动生成数据集”等流转审批规则。"],
    ["V1.0", "功能需求", "任务派发", "批量完成与前置逻辑校验：支持选中多个任务进行批量完成确认；前置逻辑校验：若选中的任务实例中存在“实际采集量低于计划采集量”且“未填写未完成补充备注”的情况，需阻断并提示错误，直至补全备注。"],
    ["V1.0", "功能需求", "任务派发", "任务新建与模式向导：提供基于模版创建及手动新建任务流程，支持配置基础参数（项目归属、任务用途、场景与子场景分类）、设备与模式选择（设备类型、实例SN、遥操类型）、采集计划量与人员分配。"],
    ["V1.0", "功能需求", "任务派发", "预设SOP动作序列编排：新建/编辑任务时支持编排动作步骤（执行末端类型、原子技能、操作对象、操作目标、默认帧数区间），支持引入动作模版导入及自然语言步骤批量解析转换。"],
    ["V1.0", "交互规则", "任务派发", "任务列表卡片及表格操作项：提供查看详情、编辑配置、复制任务、删除任务等快速操作，表格固定左侧任务ID/名称及右侧操作列，支持响应式横向滚动。"],

    # 标注审核
    ["V1.0", "功能需求", "标注审核", "标注任务列表与多维检索：支持按一级项目、任务书、任务名称、任务ID/实例ID、标注类型（框/点/范围/混合）、任务状态、标注员、审核员联合查询，并支持按“全部/待分配/进行中/已完成/暂停”分类页签进行过滤。"],
    ["V1.0", "功能需求", "标注审核", "关联数据资产新建标注任务：支持从已采集未标注的数据资产目录关联数据源，自动带出关联项目、任务书、设备类型、数据量及已编排动作步骤，快捷生成标注审核任务。"],
    ["V1.0", "功能需求", "标注审核", "人员分配与重新分配：支持选中单条或批量勾选任务，进行标注员、审核员、质检员的派发与更新，分配完成后自动更新任务状态为“进行中”或“待分配”。"],
    ["V1.0", "功能需求", "标注审核", "标注审核工作台集成：点击列表“进入”按钮跳转至标注审核工作台，支持在线标注帧区间、可视化预览动作轨迹、批注与审核打回/通过流转。"],
    ["V1.0", "交互规则", "标注审核", "标注/质检/审核多维进度可视化：列表实时渲染质检进度条、标注完成数/总数（如 150/150）、审核完成数/总数及设备SN复制功能，清晰呈现数据处理全链路状态。"],
    ["V1.0", "数据逻辑", "标注审核", "生成标注模版能力：在标注工作台完成某条具身动作数据的帧区间标定后，可一键将该数据的帧区间规则固化封存并导出为全局“标注模版”。"],

    # 模板中心
    ["V1.0", "功能需求", "模板中心", "模板中心分类管理：包含“任务模版”、“动作模版”、“标注模版”三大维度页签展示，支持模版检索、按场景数据类型筛选及快速创建入口。"],
    ["V1.0", "功能需求", "模板中心", "任务模版创建与应用：支持自定义新建任务模版，配置模版名称、模版编码、适配设备、默认采集模式及描述；支持在模版列表一键“在任务中心使用”，快速套用至任务派发流程。"],
    ["V1.0", "功能需求", "模板中心", "动作模版新建与双模式编排：支持新建动作模版，提供“结构化步骤”下拉构建（执行末端、原子技能、操作对象、操作目标、默认帧数区间）和“自然语言描述”文本批量解析转换两种编排模式。新增步骤时自动基于上一步的结束帧自动推算并填充默认帧数区间（如0-30帧、30-60帧）。"],
    ["V1.0", "功能需求", "模板中心", "动作步骤编排链条与原子规则绑定：支持在任务模版中导入已有动作模版步骤，支持动作步骤排序拖拽、动态增删，保存时自动校验动作逻辑合理性并关联物理阈值校验规则。"],
    ["V1.0", "功能需求", "模板中心", "标注模版沉淀与管理：集中展示从标注工作台固化生成的动作帧区间标注模版（如餐盘整理、工业打包等步骤与帧区间映射），支持预览封存数据及模版删除管理。"]
]

for row in data:
    ws.append(row)

# Styling
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

data_font = Font(name="Calibri", size=10)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Apply header styling
for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = align_center
    cell.border = thin_border
ws.row_dimensions[1].height = 28

# Apply data styling
for row_num in range(2, len(data) + 2):
    ws.row_dimensions[row_num].height = 45
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = data_font
        cell.border = thin_border
        if col_num in [1, 2, 3]:
            cell.alignment = align_center
        else:
            cell.alignment = align_left

# Set column widths
column_widths = {
    "A": 12,  # 版本
    "B": 14,  # 类型
    "C": 16,  # 菜单
    "D": 85   # 任务描述
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

output_path = "/Users/zhangxiaozhang/Desktop/具身智能平台原型/0714/prototype/具身智能平台_任务派发_标注审核_模板中心_需求任务表.xlsx"
wb.save(output_path)
print(f"Successfully generated Excel file at: {output_path}")
